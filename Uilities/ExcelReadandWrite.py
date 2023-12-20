import openpyxl

def getRow(file, sheetName):
    path = openpyxl.load_workbook(file)
    sheet = path[sheetName]
    return sheet.max_row
def getColumn (file, sheetName):
    path = openpyxl.load_workbook(file)
    sheet = path[sheetName]
    return sheet.max_column
def getCell(file, sheetName, r, c):
    path = openpyxl.load_workbook(file)
    sheet = path[sheetName]
    return sheet.cell(row=r,column=c).value
def getwriteCell(file, sheetName, r, c, string):
    path = openpyxl.load_workbook(file)
    sheet = path[sheetName]
    sheet.cell(row=r,column=c).value = string
    path.save(file)

def getwriteCell2(file, sheetName, r, c, intnum):
    path = openpyxl.load_workbook(file)
    sheet = path[sheetName]
    sheet.cell(row=r,column=c).value = intnum
    path.save(file)